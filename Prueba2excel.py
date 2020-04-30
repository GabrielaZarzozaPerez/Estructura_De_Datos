import openpyxl
import numpy as np
import pandas as pd
import random

libro = openpyxl.load_workbook("Lista.xlsx")
hoja = libro["Calificaciones"]

sep = "\n" + ("*" * 20) + "\n"

Continuar = True

k = 2

while Continuar == True:
    if k == 32:
        Continuar = False
    else:
        j = 2
        for i in range(5):
            print(f"\nAlumno {k-1}: {hoja.cell(row=k, column=1).value}\nMateria: {hoja.cell(row=1, column=j).value}\nCalificacion:")
            Calificacion = random.randrange(0,100)
            hoja.cell(row=k, column=j).value = Calificacion
            j = j + 1
        libro.save("Lista.xlsx")
        print(sep)
        k = k + 1
        
print("Se completo el registro de los 30 alumnos")

data = pd.read_excel("Lista.xlsx")


print(data.mean(axis=0))



